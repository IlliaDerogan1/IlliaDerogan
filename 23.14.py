import docx
import openpyxl
from collections import defaultdict

def create_xlsx(filename, *sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, table in sheets:
        ws = wb.create_sheet(name)
        for row in table:
            ws.append(row)

    wb.save(filename)

def create_sheet(price , date):
    wb = openpyxl.load_workbook(date)
    pmax = 0
    P = []
    R = []
    S = 0
    s = 0
    k = 0
    ws = wb["bills"]
    for row in ws:
        if  row[1].value == price:
            s_id , p_id , price , term = [c.value for c in row ]
            P.append(price)
            if price >= pmax :
                pmax = price
            print(s_id, p_id, price, term)
            break
    else:
        raise RuntimeError

    rmax = 0
    ws = wb["provider"]
    for row in ws:
        if row[0].value == s_id:
            s_id , name , rating , adress = [c.value for c in row]
            R.append(rating)
            if rating >= rmax :
                rmax = rating
            print(s_id , name , rating , adress)
            break
    else:
        raise RuntimeError

    print(dict)

    ws = wb["production"]
    worklist = defaultdict()
    for row in ws:
        if row[0].value in dict[id_]:
            p0 , products = [c.value for c in row]
            print(p0 , products)
            worklist[p0] = products

    for i in len(R):
        s = 0.5(P[i]/pmax)+0.5(R[i]/rmax)
        if s > S:
            S = s
            k = i
    print("Тендер виграє постачальник №",k)
    
if __name__ =="__main__":

    provider = ("provider",(("id" , "name" , "rating" , "adress"),
                  ("c01", "Доміно","4","domino@com.ua"),
                  ("c02", "Кондор", "6", "condor@com.ua")))

    production = ("production",(("id" , "name"),
             ("p01", "олівець"),
            ("p02", "ручка кулькова")))

    bills = ("bills",(("s_id","p_id","price","term"),
            ("s01","p01","2.5","5"),
            ("s02","p02","2.4","6",)))

    create_xlsx("data.xlsx", provider, production, bills)
    create_sheet("2.5","data.xlsx" )
